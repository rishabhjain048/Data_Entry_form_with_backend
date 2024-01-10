from tkinter import*
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib


root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

file=pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active 
    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['c1']="Age"
    sheet['D1']="gender"
    sheet['E1']="Address"

    file.save('Backend_data.xlsx')


def Submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)


    file=openpyxl.load_workbook('Backend_Data.xlsx')
    sheet=file.active
    sheet.Cell(column=1,row=sheet.max_row+1,value=name)
    sheet.Cell(column=2,row=sheet.max_row,value=contact)
    sheet.Cell(column=3,row=sheet.max_row,value=age)
    sheet.Cell(column=4,row=sheet.max_row,value=gender)
    sheet.Cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Backend_data.xlsx')

    messagebox.showinfo('info','detail added!')
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)

def Clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)


def Submit():
    pass

#icon
#icon_image=PhotoImage(file="meri family.jpeg")
#root.iconphoto(False,icon_image)

#heading
Label(root,text="Please fill out htese entry",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=50)


#label
Label(root,text="name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="contact No",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="Address",font=23,bg="#236273",fg="#fff").place(x=50,y=250)

#Entry
nameValue=StringVar()
contactValue=StringVar()
AgeValue=StringVar()


nameEntry=Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactEntry=Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry=Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

#gender
gender_combobox=Combobox(root,values=['Male','Female'],font='arial 14',state='r',width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Male')

addressEntry =Text(root,width=50,height=4,bd=2)


nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

Button(root,text="Submit",bg="#326273",fg="#fff",width=15,height=2,command=Submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="#fff",width=15,height=2,command=Clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="#fff",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)

root.mainloop()